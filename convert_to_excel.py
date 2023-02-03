"""
    convert_to_excel.py
    Start Date: 20230125
    Version: 1.1
    Written by: Drew Gillie

    This code is written for the purpose of completing a coding challenge for
    Laika for the Media Systems Pipeline Technical Director Position

    Version changelog:
        1.0 Original submitted code.
        1.1 Modify to run and process rather than script base

    1.1 Notes: Oh! Right, I'm writing this in python, not MEL. I can run main.

    ##########################################################################
    Challenge #1

    Write a script that exports the data in a Python dictionary to Microsoft
    Excel in spreadsheet format. The dictionary is the result of a database
    query and is provided to you in pickle file called'test_data.pkl The
    spreadsheet should include columns for:
        - Task Name,
        - Set Part Name,
        - Parent Build Name,
        - Start Date, and
        - End Date.
    The data should be sorted by earliest Start Date first.

    If a Due Date has already passed, shade the entire row red in Excel.

    You'll need to install and use an external Python library to handle the
    Excel import. We use `openpyxl` for this at Laika.

    The database returns unfriendly key names for some fields. We map these
    keys to the above listed human-readable fields in the following way:
        - Task Name = `content`
        - Set Part = `entity`
        - Parent Build = `sg_parent_build`
        - Start Date = `start_date`
        - End Date = `due_date`
    ##########################################################################


    This program will convert data from a pickle file into an excel file.
    Only relevant data is brought over and sorted from earliest date

    Overall thought process
        1. Load the data from pickle
        2. Sort data by date
        3. Create an excel file
        4. Add headers
        5. Bring data in, sorted by date
        6. Highlight past dates in red

"""
import pickle
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import sys
import os
##############################################################################



##############################################################################
""" 1. Defining the data class. """


class CapturedData:
    def __init__(self, object_name, task_name, set_part,
                 parent_build, start_date,
                 end_date):

        self.object_name = object_name
        self.task_name = task_name
        self.set_part = set_part
        self.parent_build = parent_build
        self.start_date = start_date
        self.end_date = end_date

    def print_status(self):
        print(F'object_name = {self.object_name}')
        print(F'task_name = {self.task_name}')
        print(F'set_part = {self.set_part}')
        print(F'parent_build = {self.parent_build}')
        print(F'start_date = {self.start_date}')
        print(F'end_date = {self.end_date}')

    # These two functions are only used if the data parsed, does not work.
    def set_start_date(self):
        date_format = "%Y-%m-%d"
        self.start_date = datetime.strptime(self.start_date, date_format)
        self.start_date = self.start_date.strftime("%Y-%m-%d")

    def set_end_date(self):
        date_format = "%Y-%m-%d"
        self.end_date = datetime.strptime(self.end_date, date_format)
        self.end_date = self.end_date.strftime("%Y-%m-%d")

##############################################################################
""" 2. Create an excel file. """

def check_file_type(file_to_check):
    if file_to_check.endswith('.pkl'):
        return True
    else:
        return False


def set_cell_value(worksheet, row, col, value):
    cell = worksheet.cell(row=row, column=col)
    if not isinstance(value, str):
        # converted_value = ''.join(value)
        converted_value = str(value)
    else:
        converted_value = value
        
    cell.value = converted_value


##############################################################################
""" 3. Add data to excel file after fixing some data. """

"""
This section of code is all about taking the input data from the .pkl
file and converting it into simple lists.  The lists will then be added to
excel.  Some issues will arise as some of the input data may return None.
"""


def remove_brackets(input):
    return str(input).strip('[]')


def remove_quote(input):
    return str(input).strip(" '' ")


def catch_key(catch_list):  # Catch None error and return 'Not acailable'.
    """
    Input data = [list, key, nested key].
    The list that is passed into this function contains 3 items.  The first
    two variables are on each item that passes through, however the
    third nested key variable only comes in with a few pieces of data.
    Therefore, if it exists, run the case where all three are tested.
    The thing being tested in this function is if a piece of data from
    the .pkl file did not have an entry.  If no entry, then return a
    phrase 'Not available' in order to keep the code running and note
    this on the excel worksheet.
    """
    data_list = catch_list[0]
    key_catch = catch_list[1]

    if len(catch_list) == 3:
        nested_catch = catch_list[2]
        if key_catch in data_list:
            if data_list[key_catch] is not None:
                if nested_catch in data_list[key_catch]:
                    return data_list[key_catch][nested_catch]
                else:
                    return 'Not available'
            else:
                return 'Not available'
        else:
            return 'Not available'
    else:
        if key_catch in data_list:
            return data_list[key_catch]
        else:
            return 'Not available'


def get_value(input_data, key_code, title_list):  # Find specific key-values.
    """
    This is the main data conversion function.

    The goal is to find the requested data that should be on the
    spreadsheet.

    When an object is created, it will set object traits to the value it
    should be.  This function retreives the value for the associated key.

    Some pieces of data may need to be corrected, this is done by removing
    brackts "[]" and single quotes around text " '' ".
    """

    switcher = {
        title_list[0]: ('content', None),
        title_list[1]: ('entity', 'code'),
        title_list[2]: ('sg_parent_build', 'code'),
        title_list[3]: ('start_date', None),
        title_list[4]: ('due_date', None)
    }
    if key_code not in switcher:
        return 'Invalid Key (key_code'

    # # Find the corresponding key-value pair
    key_values = switcher.get(key_code)  # Get the key-cade value

    # if the pair has a nested list, get the nested value
    if key_values[1] is not None:
        content_list = [input_data, key_values[0], key_values[1]]

        # returned_key = input_data[key_values[0]][key_values[1]]
    else:  # get the normal value
        content_list = [input_data, key_values[0]]

        # returned_key = input_data[key_values[0]]
    returned_key = catch_key(content_list)

    # Some data may not exist,
    if returned_key is None:
        returned_key = 'Not available'

    returned_key = remove_brackets(returned_key)
    returned_key = remove_quote(returned_key)

    return returned_key


def process_data(data_in, data_categories):
    print('####################------##############')
    # Convert pickled data into a list.
    unpickled_data = open(data_in, 'rb')
    data = pickle.load(unpickled_data)

    # Sort data by start date now
    sorted_data = sorted(data, key=lambda x: x['start_date'])

    # Find size - how many lists are inside this list.
    list_size = len(sorted_data)
    print('Size of list = {}\n'.format(list_size))

    # List of objects created from data.
    captured_data_list = []
    for i in range(0, list_size):
        obj_name = ('object_row_' + str(i+2))
        captured_data_list.append(CapturedData(obj_name, '', '', '', '', ''))

    for index, objects in enumerate(captured_data_list):
        """
        For each object, set specific_data to list index
        The List index is which data entry point in the larger list.  Then set
        the object attribute to the key:value pair
        """
        task_header = str(data_categories[0])
        set_part_header = str(data_categories[1])
        parent_build_header = str(data_categories[2])
        start_date_header = str(data_categories[3])
        end_date_header = str(data_categories[4])

        row_number = index + 2
        print(F'index {row_number}')
        specific_data = sorted_data[index]
        objects.task_name = get_value(specific_data, task_header,
                                      data_categories)
        objects.set_part = get_value(specific_data, set_part_header,
                                     data_categories)
        objects.parent_build = get_value(specific_data, parent_build_header,
                                         data_categories)
        objects.start_date = get_value(specific_data, start_date_header,
                                       data_categories)
        objects.end_date = get_value(specific_data, end_date_header,
                                     data_categories)

        objects.set_start_date()
        objects.set_end_date()
        

        set_cell_value(ws, row_number, 1, objects.task_name)
        set_cell_value(ws, row_number, 2, objects.set_part)
        set_cell_value(ws, row_number, 3, objects.parent_build)
        set_cell_value(ws, row_number, 4, objects.start_date)
        set_cell_value(ws, row_number, 5, objects.end_date)

        objects.print_status()
        print(F'------------------------------------\n')


##############################################################################
# 5. Format data
""" This section of code is dedicated to reformatting the data on excel. """


def fill_row(selected_row):  # Give a row number to fill in red.
    row_cells = ws[selected_row]  # Select row from input.

    # Select the color red.
    fill_color = PatternFill(start_color='FFC7CE',
                             end_color='FFC7CE', fill_type='solid')
    font_change = Font(name='Arial', size=12, color='8B0000')

    # Set the fill color of all of the cells in the row.
    for cell in row_cells:
        cell.fill = fill_color
        cell.font = font_change


def date_to_check_from_excel():  # checks date, colors row red.
    # Find the column number of the "End Date" column
    end_date_col = None
    for row in ws.iter_rows(values_only=True):
        if 'End Date' in row:
            end_date_col = row.index('End Date') + 1
            break

    # Check the date values in the "End Date" column
    if end_date_col:
        for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
            try:
                date_to_check = datetime.strptime(row[end_date_col-1], '%Y-%m-%d')
                if date_to_check < datetime.now():
                    date_to_check = date_to_check.strftime("%Y-%m-%d")
                    print(f'Row {row_num} has an End Date that has already'
                          f' passed: {date_to_check}')
                    fill_row(row_num)
            except ValueError as e:
                if row_num == 1:
                    continue
                else:
                    print(f'Error parsing date in row {row_num}: {e}')
    else:
        print('Could not find End Date column')


def format_title(column_headers):  # Format the titles / headers.
    # Set the Title font.
    title_font = Font(name='Arial', size=20, color='ffffff')

    # Set Title fill.
    title_fill = PatternFill(start_color='000000', end_color='000000',
                             fill_type='solid')

    # Iterate through column_names.
    for i, title in enumerate(column_headers):
        column_letter = get_column_letter(i + 1)  # Get the column letter.
        column_number = column_letter + '1'

        # Apply the formatting.
        ws[column_number].font = title_font
        ws[column_number].fill = title_fill


def center_align_cells(worksheet):  # Alaign data to center.
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')


def right_align_column_a():  # Aligns first column data to right.
    for row in ws.iter_rows(min_row=1, max_col=1):
        for cell in row:
            if cell.row != 1:
                cell.alignment = Alignment(horizontal='right',
                                           vertical='center')


def autofit_columns(column_number):  # Adjust column width based on max cell.
    max_size = 0  # Initialize the max size.

    # Iterate through cells in the column.
    for cell in ws[column_number]:
        # Get the length of the cell value and font size.
        length = len(str(cell.value))
        font_size = cell.font.size
        max_size = max(max_size, length + font_size)

    average_size = 15  # Guess and check to see how it looks.
    new_column_size = ((average_size + max_size) / 2) - 2

    # print(max_size)  # Only for testing purposes---.
    ws.column_dimensions[column_number].width = new_column_size


def check_unavailable_data():  # Search for data with no entry point.
    # Set the destination font change.
    font_change = Font(name='Arial', size=12, color='8B0000',
                       bold=True, italic=True)

    """
    Search all of the data to find 'Not available'.
    This strange error is documented, but all instances of 'Not available'
    must be a lower case 'a'.

    When 'Not available' is found, set the font change.
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 'Not available':
                # Apply the formatting.
                cell.font = font_change


def format_data(column_headers):
    format_title(column_headers)  # Change Headers on columns.
    date_to_check_from_excel()  # Color row red if date has passed.
    center_align_cells(ws)  # Center all data.
    right_align_column_a()  # Make task name adjusted to the right.

    columns = ['A', 'B', 'C', 'D', 'E']
    for column in columns:
        autofit_columns(column)  # Scale column width.

    check_unavailable_data()  # Highlight the missing data.

##############################################################################
##############################################################################


if __name__ == '__main__':
    """
    Script Outline.
    1. Check if file is a valid file type.
    2. Create an excel file and fill in the titles.
    3. Bring in the file that runs with the script to convert.
    4. Process data.
    5. Format data.
    6. Save File
    """
    if len(sys.argv) > 1:
        if check_file_type(sys.argv[1]):
            print('\nStarting Script\n')  # Only for testing purposes--------.
        else:
            print(F'This file type cannot be converted with this script.')
    else:
        print('\n======================================================')
        print('There was no file passed.\nPlease run the command like:')
        print('python3 {file_path_name} {file_to_process_path}')
        print('======================================================\n')


    """ Variables to change for different type of data sheets. """
    column_names = ['Task', 'Set Part', 'Parent Build', 'Start Date',
                    'End Date']
    
    # 1. Create an excel file using the file name of the data to be processed.
    wb = Workbook()  # Create a workbook.
    ws = wb.active  # Add worksheet.
    ws.title = 'Formatted'  # Change title.

    # Pass titles from column_names list intto excel.
    for i in range(1, (len(column_names)+1)):
        title = str(column_names[i-1])

        set_cell_value(ws, 1, i, title)

    
    # Input file when script is run through command line
    input_file = sys.argv[1]  # 'test_data.pkl'

    # Unpickle, sort, and set up excel file.
    process_data(input_file, column_names)

    # Select data in excel and format it.
    format_data(column_names)

    # Create file name and save file.
    output_file = (input_file + '_converted_data.xlsx')
    wb.save(output_file)
    print('\nFinished Processing')  # Only for testing purposes----------.
